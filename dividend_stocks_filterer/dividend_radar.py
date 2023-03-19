import requests
import openpyxl
from bs4 import BeautifulSoup
from retrying import retry
from cachetools import cached, TTLCache


class DividendRadar:
    def __init__(self, dividend_radar_url: str, local_file: str) -> None:
        """
        This class handles everything to do with the dividend radar file, finding, downloading and version checking

        :param dividend_radar_url: the url of the page with the dividend radar download button
        :param local_file: where to save the local dividend radar xlsx file
        """
        self.dividend_radar_url = dividend_radar_url
        self.local_file = local_file
        self.latest_version_url = None
        self.latest_version = None
        self.latest_local_version = None

    @cached(cache=TTLCache(maxsize=1024, ttl=3600))
    @retry(wait_exponential_multiplier=2500, wait_exponential_max=10000, stop_max_attempt_number=10)
    def find_latest_version(self) -> str:
        """
        Finds the latest version of the dividend radar xlsx file, note I'm wrapping this with retries to avoid network
        glitches and then caching the result as a quick way to avoid spamming the site

        :return: the latest version of the file
        """
        page = requests.get(self.dividend_radar_url)
        soup = BeautifulSoup(page.content, "html.parser")
        self.latest_version_url = soup.find(class_="link-block w-inline-block").attrs['href']
        self.latest_version = self.latest_version_url[-15:-5]
        return self.latest_version

    def check_if_local_is_latest(self) -> bool:
        """
        Checks if the local dividend radar file version is the latest one

        :return: True if the local version is the latest version of the file available, False otherwise
        """
        try:
            if self.latest_local_version == self.find_latest_version():
                return True
            else:
                return False
        except FileNotFoundError:
            return False

    def download_latest_version(self) -> None:
        """
        Gets the latest version of the dividend radar file and puts it in the local path declared in the function
        """
        self.find_latest_version()
        latest_radar_file = requests.get(self.latest_version_url)
        with open(self.local_file, 'wb') as f:
            f.write(latest_radar_file.content)
        self.latest_local_version = self.latest_version

    @cached(cache=TTLCache(maxsize=1024, ttl=60))
    def read_radar_file_to_dict(self) -> dict:
        """

        :return dividend_radar_dict: A dict of the relevant data from the file
        """
        radar_dict = {}
        wb = openpyxl.load_workbook(self.local_file)
        sheet = wb["All"]

        for row in sheet.iter_rows(min_row=4):
            radar_dict[row[0].value] = {}
            for cell in row:
                radar_dict[row[0].value][sheet.cell(row=3, column=cell.column).value] = cell.value
        return radar_dict
